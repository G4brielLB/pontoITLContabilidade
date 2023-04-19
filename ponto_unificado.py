import tkinter as tk
from datetime import datetime
import telebot
import openpyxl
from openpyxl.styles import Font
# Recebe as funções para a planilha do excel, modularizado em outro arquivo
#from planilha_ponto import verificar_data, atualizar_planilha, criar_aba, totalizar

# Configurações do bot do Telegram
API_KEY = '6049860098:AAGj7yTazEXn4yUFfcNJ44dh0pgJ4VYnOvA'
CHAT_ID = '-927904484'
NOMES_PERMITIDOS = ['Fernando', 'Isadora', 'Neto', 'Karina', 'Gabriel']
# Coloca em ordem alfabética
NOMES_PERMITIDOS.sort()
# COLUNAS EXCEL
letras = ["B", "C", "D", "E", "F"]
# Pega a data atual
data = datetime.now().strftime("%d-%m-%Y")

# Verifica se a data já está na planilha, ou é um novo dia
def verificar_data(datas):
    # Se a aba da data referida já existir, retorna True, se ainda não tiver criado a aba, retorna False
    if data not in datas:
        return False
    else:
        return True
    
# Cria uma nova aba com uma nova data
def criar_aba(arquivo, data, firstTime=False):
    # Verifica se é a primeira vez que está criando o arquivo pela primeira vez
    if firstTime:
        # Se está pela primeira vez, seleciona a primeira aba ativa
        planilha = arquivo.active

    # Se não estiver criando pela primeira vez, cria uma aba e muda para a mesma
    else:
        # Cria a aba
        arquivo.create_sheet(data)
        # Seleciona a aba
        planilha = arquivo[data]
    # Renomeia a aba
    planilha.title = data
    # Cria uma variável para a fonte em negrito
    fonte_negrito = Font(bold=True)
    # Coloca a primeira célula com a data e coloca em negrito
    planilha.cell(row=1, column=1, value=data)
    planilha['A1'].font = fonte_negrito
    # Renomeia a primeira célula de cada coluna
    planilha['B1'].value = "ENTRADA 1"
    planilha['C1'].value = "SAÍDA 1"
    planilha['D1'].value = "ENTRADA 2"
    planilha['E1'].value = "SAÍDA 2"
    planilha['F1'].value = "TOTAL"
    # Faz um loop em todas as colunas e coloca em negrito
    for a in letras: 
        planilha[f'{a}1'].font = fonte_negrito
    # Faz um loop nas linhas e coloca o nome dos funcionários na primeira célula de cada linha
    for i, pessoa in enumerate(NOMES_PERMITIDOS):
        planilha.cell(row=i+2, column=1, value = pessoa)

# Atualiza a planilha com um novo horário (input)
def atualizar_planilha(funcionario, horario, planilha):
    # Variáveis que recebem o número de linhas e colunas 
    linha_max = planilha.max_row
    coluna_max = planilha.max_column
    # Faz um loop nas linhas dos funcionários
    for i in range(2, linha_max + 1):
        # Quando a célula da planilha for o nome do funcionário do input 
        if planilha.cell(row=i, column=1).value == funcionario:
            # j é a coluna, a partir do 2, pois o 1 é o "index"
            j = 2
            # Enquanto o j (coluna) não chegar no final das colunas
            while j <= coluna_max:
                # Seleciona a celula
                celula = planilha.cell(row=i, column=j)
                # Se a célula estiver vazia, vai adicionar nela
                if celula.value == None:
                    # Transforma a celula de formato de string para o formato de horario (numero)
                    celula.number_format = 'hh:mm:ss'
                    # Coloca o horário de registro na célula
                    celula.value = horario
                    # Pega o status (entrada ou saída) a partir do título da coluna
                    status = planilha.cell(row=1, column=j).value
                    # Se o título da coluna for ENTRADA, retorna que o funcionário entrou
                    if "ENTRADA" in status:
                        return(f"{funcionario} entrou às {horario}")
                    # Se o título da coluna for SAÍDA, retorna que o funcionário entrou
                    elif "SAÍDA" in status:
                        return(f"{funcionario} saiu às {horario}")
                    # Se achou a célula vazia, e já adicionou, ele da break no loop
                    break
                # Continua o loop, indo para a próxima coluna
                j += 1

# Cria a função total na última coluna da planilha
def totalizar(planilha):
    # Faz um loop na coluna TOTAL
    for i in range(2, planilha.max_row + 1):
        # Coloca a fórmula dentro da célula
        planilha[f'F{i}'] = f'=C{i}-B{i}+E{i}-D{i}'
        # Transforma o formato da célula em número de horas
        planilha[f'F{i}'].number_format = 'hh:mm:ss'
    
# Função para enviar mensagem para o Telegram
def enviar_telegram(mensagem):
    bot = telebot.TeleBot(token=API_KEY)
    bot.send_message(chat_id=CHAT_ID, text=mensagem)

def verificar_nome(nome):
    # Verifica se o nome realmente está na lista
    if nome in NOMES_PERMITIDOS:
        return True
    else:
        return False

# Função para registrar a entrada/saída do colaborador
def registrar_ponto():
    nome = var_nome.get()
    # Verifica se realmente possui um nome permitido
    if verificar_nome(nome):
        hora_atual = datetime.now().strftime("%H:%M:%S")
        data_atual = datetime.now().strftime(" - %d/%m/%Y")
        # Tenta abrir a planilha do excel (se já existente)
        try:
            arquivo = openpyxl.load_workbook("Ponto ITL Contabilidade.xlsx")
            datas = arquivo.sheetnames

            if verificar_data(datas):
                # Atualização normal 
                mensagem = atualizar_planilha(nome, hora_atual, arquivo[data])
                print(mensagem)

            else:
                criar_aba(arquivo, data)
                # Ao criar uma nova aba em um novo dia, esse funcionário e esse horário serão adicionados
                mensagem = atualizar_planilha(nome, hora_atual, arquivo[data])
                print(mensagem)

            arquivo.save("Ponto ITL Contabilidade.xlsx")

        # Cria a planilha do excel com algumas configurações
        except:
            arquivo = openpyxl.Workbook()
            criar_aba(arquivo, data, True)
            planilha = arquivo[data]
            # Ao criar o arquivo, esse funcionário e esse horário serão adicionados
            mensagem = atualizar_planilha(nome, hora_atual, planilha)
            print(mensagem)
            totalizar(planilha)
            arquivo.save("Ponto ITL Contabilidade.xlsx")

        #mensagem = f"{nome} registrou ponto às {hora_atual}"

        enviar_telegram(mensagem + data_atual)
        label_mensagem.configure(text=mensagem)
        # Reseta a opção para a inicial
        var_nome.set('----------')
    else:
        mensagem = "Selecione um nome válido!"
        label_mensagem.configure(text=mensagem)

# Função para atualizar o relógio
def atualizar_relogio():
    data_hora_atual = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    label_data_hora.configure(text=data_hora_atual)
    root.after(1000, atualizar_relogio)

# Criação da janela principal
root = tk.Tk()
root.title("Controle de Ponto")

# Criação dos widgets
label_nome = tk.Label(root, text="Nome:")
var_nome = tk.StringVar(value="----------")
option_menu_nome = tk.OptionMenu(root, var_nome, *NOMES_PERMITIDOS)
button_registrar = tk.Button(root, text="Registrar", command=registrar_ponto)
label_mensagem = tk.Label(root, text="")
label_data_hora = tk.Label(root, font=('Arial', 18), pady=10)



# Posicionamento dos widgets na janela
label_nome.grid(row=0, column=0)
option_menu_nome.grid(row=0, column=1)
button_registrar.grid(row=1, column=0, columnspan=2)
label_mensagem.grid(row=2, column=0, columnspan=2)
label_data_hora.grid(row=3, column=0, columnspan=2)

# Atualização do relógio
atualizar_relogio()

# Inicialização da janela
root.mainloop()